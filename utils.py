import os
from datetime import date
from models.models import Auditoria
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet
from typing import cast

def build_seguimiento_workbook(year: int, auditorias: list[Auditoria]) -> Workbook:
    wb = Workbook()
    ws = cast(Worksheet, wb.active)
    ws.title = "SEGUIMIENTO MEMORANDOS"

    col_widths = {
        "A": 10,
        "B": 14,
        "C": 11,
        "D": 25,
        "E": 30,
        "F": 90,
        "G": 9,
        "H": 12,
    }
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    logo_path = os.path.join(os.path.dirname(__file__), "..", "assets", "logo.png")
    if os.path.exists(logo_path):
        img = Image(logo_path)
        img.width = 180
        img.height = 60
        ws.add_image(img, "A1")

    ws.merge_cells("A3:H3")
    ws["A3"] = (
        "CONTROL Y SEGUIMIENTO A INFORMES, MEMORANDOS EMITIDOS Y RESPUESTAS RECIBIDAS"
    )
    ws["A3"].font = Font(bold=True, size=12)
    ws["A3"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A6:H6")
    ws["A6"] = f"AÑO: {year}"
    ws["A6"].font = Font(bold=True, size=11)
    ws["A6"].alignment = Alignment(horizontal="center", vertical="center")

    header_fill = PatternFill("solid", fgColor="D9D9D9")
    header_font = Font(bold=True, size=10)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.merge_cells("A8:A9")
    ws.merge_cells("B8:B9")
    ws.merge_cells("C8:C9")
    ws.merge_cells("D8:D9")
    ws.merge_cells("E8:E9")
    ws.merge_cells("F8:F9")
    ws.merge_cells("G8:H8")

    ws["G8"] = "Respuesta"
    ws["G8"].fill = header_fill
    ws["G8"].font = header_font
    ws["G8"].alignment = center

    headers = {
        "A8": "Fecha",
        "B8": "Consecutivo DAI",
        "C8": "No. De Radicación OnBase",
        "D8": "Destinatario",
        "E8": "Tema",
        "F8": "Observaciones y/o Recomendaciones",
        "G9": "Fecha",
        "H9": "No. De Radicación OnBase",
    }

    for cell, value in headers.items():
        ws[cell] = value
        ws[cell].fill = header_fill
        ws[cell].font = header_font
        ws[cell].alignment = center

    ws.row_dimensions[8].height = 14
    ws.row_dimensions[9].height = 40

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows(min_row=8, max_row=9, min_col=1, max_col=8):
        for cell in row:
            cell.border = border

    consecutive_map: dict[int, int] = {}
    for index, auditoria in enumerate(auditorias, start=1):
        auditoria_id = cast(int, auditoria.id_aud)
        consecutive_map[auditoria_id] = index

    row_index = 10
    for auditoria in auditorias:
        auditoria_id = cast(int, auditoria.id_aud)
        consecutive = consecutive_map[auditoria_id]
        if not auditoria.mejoras:
            continue
        for mejora in auditoria.mejoras:
            date_onbase = cast(date, auditoria.date_onbase)
            radicate_onbase = cast(str, auditoria.radicate_onbase)
            area = cast(str, auditoria.area)
            topic = cast(str, auditoria.topic)
            description = cast(str, mejora.description)
            ws.cell(row=row_index, column=1, value=date_onbase)
            ws.cell(row=row_index, column=2, value=consecutive)
            ws.cell(row=row_index, column=3, value=radicate_onbase)
            ws.cell(row=row_index, column=4, value=area)
            ws.cell(row=row_index, column=5, value=topic)
            ws.cell(row=row_index, column=6, value=description)
            ws.cell(row=row_index, column=7, value=None)
            ws.cell(row=row_index, column=8, value=None)

            for col_idx in range(1, 9):
                cell = ws.cell(row=row_index, column=col_idx)
                cell.border = border
                cell.alignment = Alignment(
                    horizontal="center" if col_idx != 6 else "left",
                    vertical="top",
                    wrap_text=col_idx == 6,
                )
            row_index += 1

    return wb